"""
평가파일 종합 점검 Worker

[검사 항목]
  1. 변동 시트 존재 여부 + B1 값 + B1와 파일명 prefix 일치 여부
  2. 대장MMDD 시트 존재 여부 + 미지급분배금 키워드 존재 여부
  3. Named Range F_1 위치 → 아래3 오른쪽1 → 펀드코드(Fxxxxx) 검증
  4. {펀드코드}_nav, {펀드코드}_inv Named Range 위치
  5. _nav 위치 → 아래4 오른쪽1 → 장부기준일(YYYY-MM-DD) vs 대장MMDD 월일 비교
  6. 02. 증빙 폴더 내 해당 분기 메일 존재 여부
  7. 01. 자료 폴더 내 파일 목록
"""

import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from .base import AbstractWorker
from utils.domain import (
    EXCEL_EXTENSIONS, CHANGE_SHEETS_NORM,
    DAEJANG_RE, PREFIX_RE, FUND_CODE_RE,
    F1_NAME, KEYWORD_MISUI,
    F1_OFFSET_ROW, F1_OFFSET_COL,
    NAV_OFFSET_ROW, NAV_OFFSET_COL,
    DATE_IN_NAME_RE,
)
from utils.formats import FMT_DATE, FMT_DATETIME, fmt_file_stamp
from utils.messages import Msg


# ── 공통 유틸 ─────────────────────────────────────────────────────────

def _extract_bracket_reason(name: str) -> str:
    m = re.match(r"^\[([^\]]+)\]", name)
    return m.group(1) if m else name


def _find_target_files(folder: Path, prefix: str, include_ga: bool = False) -> List[Path]:
    """
    폴더 바로 아래(1단계)에서 대상 파일 탐색.
    - prefix가 파일명 어디든 포함되면 대상
    - include_ga=False(기본): '가평가' 포함 파일 제외
    """
    return sorted([
        f for f in folder.iterdir()          # rglob → iterdir (1단계만)
        if f.is_file()
        and f.suffix.lower() in EXCEL_EXTENSIONS
        and prefix in f.name
        and not f.name.startswith("~$")
        and (include_ga or "가평가" not in f.name)
    ])


def _detect_prefixes(base_path: Path) -> List[str]:
    """하위 폴더 엑셀 파일명(1단계)에서 prefix 자동 감지"""
    found = set()
    for folder in base_path.iterdir():
        if not folder.is_dir() or folder.name.startswith("["):
            continue
        for f in folder.iterdir():           # rglob → iterdir
            if f.is_file() and f.suffix.lower() in EXCEL_EXTENSIONS and not f.name.startswith("~$"):
                m = PREFIX_RE.search(f.name)
                if m:
                    found.add(m.group(1).upper().replace("q", "Q"))
    return sorted(found)


def _cell_str(ws, addr: str) -> str:
    v = ws[addr].value
    return "" if v is None else str(v).strip()


def _sheet_contains_keyword(ws, keyword: str) -> bool:
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None and keyword in str(cell):
                return True
    return False


def _parse_ref(ref_str: str):
    sheet_match = re.match(r"^'?([^'!]+)'?!", ref_str)
    sheet_name  = sheet_match.group(1).strip("'") if sheet_match else None
    cell_part   = re.sub(r"^'?[^'!]+'?!", "", ref_str).replace("$", "")
    cell_match  = re.match(r"([A-Za-z]+)(\d+)", cell_part)
    if not cell_match:
        raise ValueError(f"셀 주소 파싱 실패: {ref_str}")
    col = column_index_from_string(cell_match.group(1))
    row = int(cell_match.group(2))
    return sheet_name, row, col


def _addr(sheet_name, row: int, col: int) -> str:
    cell = f"{get_column_letter(col)}{row}"
    return f"{sheet_name}!{cell}" if sheet_name else cell


def _quarter_year_months(prefix: str) -> List:
    m = re.match(r"(\d{2})\.([1-4])Q", prefix)
    if not m:
        return []
    year = 2000 + int(m.group(1))
    q    = int(m.group(2))
    MAP  = {
        1: [(year, [3, 4])],
        2: [(year, [6, 7])],
        3: [(year, [9, 10])],
        4: [(year, [12]), (year + 1, [1])],
    }
    return MAP[q]


# ── 날짜 감지 ────────────────────────────────────────────────────────

def _get_file_date(f: Path) -> datetime:
    """
    파일명에서 날짜 파싱 시도 (YYYY-MM-DD, YYYYMMDD 등).
    실패 시 수정일(mtime) 사용.
    """
    m = DATE_IN_NAME_RE.search(f.stem)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return datetime.fromtimestamp(f.stat().st_mtime)


# ── 폴더 검사 ─────────────────────────────────────────────────────────

def _get_01_filenames(folder: Path):
    target = next(
        (d for d in folder.iterdir() if d.is_dir() and d.name.startswith("01. 자료")),
        None
    )
    if target is None:
        return "해당없음"
    return sorted(f.name for f in target.rglob("*") if f.is_file())


def _check_02_mails(folder: Path, prefix: str) -> str:
    target = next(
        (d for d in folder.iterdir() if d.is_dir() and d.name.startswith("02. 증빙")),
        None
    )
    if target is None:
        return "❌ 폴더 없음"

    periods = _quarter_year_months(prefix)
    if not periods:
        return "⚠️ 분기파싱 실패"

    count = 0
    for f in target.rglob("*"):
        if not f.is_file() or f.name.startswith("~$"):
            continue
        dt = _get_file_date(f)   # 파일명 날짜 우선, 실패 시 mtime
        for (yr, months) in periods:
            if dt.year == yr and dt.month in months:
                count += 1
                break

    return f"✅ {count}개" if count > 0 else "❌ 없음"


# ── 파일 분석 ─────────────────────────────────────────────────────────

def _extract_file_prefix(stem: str, prefix: str) -> str:
    """
    파일명에서 변동시트 B1과 비교할 prefix 추출.
    prefix가 파일명 어디든 있을 수 있으므로 PREFIX_RE로 직접 탐지.
    예) "26.1Q_펀드A" → "26.1Q" / "펀드A_26.1Q" → "26.1Q"
    """
    m = PREFIX_RE.search(stem)
    if m:
        return m.group(1).upper().replace("q", "Q")
    # PREFIX_RE 미매칭 시 (260331 등 직접입력 케이스) → prefix 그대로 반환
    return prefix


def _get_nav_cell(ws, base_row: int, base_col: int, dr: int, dc: int):
    """ws에서 (base_row+dr, base_col+dc) 셀 값 반환 — 클로저 대신 명시적 함수"""
    return ws.cell(row=base_row + dr, column=base_col + dc).value


def _safe_ref_value(named_range) -> str:
    """
    openpyxl DefinedName.value는 다중 범위일 때 ','로 이어진 문자열 반환 가능.
    첫 번째 범위만 사용.
    """
    val = named_range.value or ""
    return val.split(",")[0].strip()


def _analyze_file(file_path: Path, prefix: str) -> Dict[str, Any]:
    file_prefix = _extract_file_prefix(file_path.stem, prefix)

    r = {
        "change_note": "", "change_b1": "", "prefix_match": "-",
        "daejang_exists": "❌", "daejang_name": "", "daejang_kw": "-",
        "f1_addr": "", "target_cell": "", "fund_code": "", "is_valid": False,
        "nav_ref": "없음", "inv_ref": "없음",
        "jangbu_cell": "", "jangbu_value": "", "date_match": "-",
        "misui_a": "", "misui_b": "", "misui_note": "-",
        "gita_a": "", "gita_b": "", "gita_note": "-",
        "error": "",
    }

    try:
        wb       = load_workbook(file_path, data_only=True)
        sh_names = wb.sheetnames
        dn       = wb.defined_names

        # 1. 변동 시트 — 공백 무시 비교
        found = [n for n in sh_names if re.sub(r'\s+', '', n) in CHANGE_SHEETS_NORM]
        if len(found) == 1:
            sn                = found[0]
            b1                = _cell_str(wb[sn], "B1")
            r["change_note"]  = f"✅ {sn}"
            r["change_b1"]    = b1
            r["prefix_match"] = "✅" if b1 == file_prefix else "❌"
        elif len(found) == 0:
            r["change_note"] = "❌ 없음"
        else:
            r["change_note"] = f"⚠️ {len(found)}개: {', '.join(found)}"

        # 2. 대장MMDD 시트
        dj = next((n for n in sh_names if DAEJANG_RE.match(n)), None)
        if dj:
            r["daejang_exists"] = "✅"
            r["daejang_name"]   = dj
            r["daejang_kw"] = "✅" if _sheet_contains_keyword(wb[dj], KEYWORD_MISUI) else "❌"

        # 3. F_1 Named Range
        f1 = dn.get(F1_NAME)
        if f1:
            try:
                sn_f1, row_f1, col_f1 = _parse_ref(_safe_ref_value(f1))  # ✅ 다중범위 대응
                r["f1_addr"] = _addr(sn_f1, row_f1, col_f1)
                ws_f1 = wb[sn_f1] if sn_f1 and sn_f1 in sh_names else wb.active

                t_row = row_f1 + F1_OFFSET_ROW
                t_col = col_f1 + F1_OFFSET_COL
                r["target_cell"] = _addr(ws_f1.title, t_row, t_col)

                raw     = ws_f1.cell(row=t_row, column=t_col).value
                val_str = str(raw).strip() if raw is not None else ""

                if FUND_CODE_RE.match(val_str):
                    r["is_valid"]  = True
                    r["fund_code"] = val_str

                    # 4. _nav / _inv
                    nav_raw_ref = None
                    for sfx in ("nav", "inv"):
                        nm = dn.get(f"{val_str}_{sfx}")
                        if nm:
                            ref_val = _safe_ref_value(nm)  # ✅ 다중범위 대응
                            r[f"{sfx}_ref"] = _addr(*_parse_ref(ref_val))
                            if sfx == "nav":
                                nav_raw_ref = ref_val
                        else:
                            r[f"{sfx}_ref"] = "없음"

                    # 5. 장부기준일 / 미수이자 / 기타자산
                    if nav_raw_ref:
                        try:
                            sn_nav, row_nav, col_nav = _parse_ref(nav_raw_ref)
                            ws_nav = wb[sn_nav] if sn_nav and sn_nav in sh_names else None
                            if ws_nav:
                                # ✅ 클로저 대신 명시적 함수 호출
                                def nv(dr, dc):
                                    return _get_nav_cell(ws_nav, row_nav, col_nav, dr, dc)

                                jb_raw = nv(NAV_OFFSET_ROW, NAV_OFFSET_COL)
                                if jb_raw is not None:
                                    jb_str = (
                                        jb_raw.strftime(FMT_DATE)
                                        if hasattr(jb_raw, "strftime")
                                        else str(jb_raw).strip()
                                    )
                                    r["jangbu_cell"]  = _addr(ws_nav.title, row_nav + NAV_OFFSET_ROW, col_nav + NAV_OFFSET_COL)
                                    r["jangbu_value"] = jb_str

                                    if r["daejang_name"]:
                                        dj_m = DAEJANG_RE.match(r["daejang_name"])
                                        jb_m = re.match(r"\d{4}-(\d{2})-(\d{2})", jb_str)
                                        if dj_m and jb_m:
                                            r["date_match"] = (
                                                "✅" if int(dj_m.group(1)) == int(jb_m.group(1))
                                                        and int(dj_m.group(2)) == int(jb_m.group(2))
                                                else "❌"
                                            )

                                mi_a, mi_b = nv(7, 1), nv(7, 2)
                                r["misui_a"] = mi_a if mi_a is not None else ""
                                r["misui_b"] = mi_b if mi_b is not None else ""
                                if mi_a is not None and mi_b is not None:
                                    try:
                                        r["misui_note"] = "✅" if float(mi_a) < float(mi_b) else "⚠️ 확인 필요"
                                    except (TypeError, ValueError):
                                        r["misui_note"] = "⚠️ 비교불가"

                                gt_a, gt_b = nv(9, 1), nv(9, 2)
                                r["gita_a"] = gt_a if gt_a is not None else ""
                                r["gita_b"] = gt_b if gt_b is not None else ""

                                gt_a_str = str(gt_a).strip() if gt_a is not None else ""
                                gt_b_str = str(gt_b).strip() if gt_b is not None else ""

                                if not gt_a_str and not gt_b_str:
                                    # 둘 다 빈칸 → 정상
                                    r["gita_note"] = "✅"
                                elif gt_a_str and gt_b_str:
                                    try:
                                        r["gita_note"] = "✅" if float(gt_a_str) == float(gt_b_str) else "❌"
                                    except (TypeError, ValueError):
                                        r["gita_note"] = "❌ 비교불가"
                                else:
                                    # 한쪽만 값이 있음
                                    r["gita_note"] = "❌"

                        except Exception as e:
                            r["error"] = r["error"] or f"NAV분석: {e}"
            except Exception as e:
                r["error"] = r["error"] or f"F_1: {e}"

        wb.close()

    except Exception as e:
        r["error"] = str(e)

    return r


# ── 엑셀 저장 ─────────────────────────────────────────────────────────

def _build_styles() -> dict:
    bd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    return {
        "hf"   : Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "bf"   : Font(name="Arial", size=10),
        "bold" : Font(name="Arial", bold=True, size=10),
        "gray" : Font(name="Arial", size=10, color="AAAAAA"),
        "c"    : Alignment(horizontal="center", vertical="center"),
        "l"    : Alignment(horizontal="left",   vertical="center"),
        "bd"   : bd,
        "h0"   : PatternFill("solid", start_color="2F5496"),
        "h1"   : PatternFill("solid", start_color="1A3A6B"),
        "h2"   : PatternFill("solid", start_color="1A5C38"),
        "h3"   : PatternFill("solid", start_color="5C1A5C"),
        "h4"   : PatternFill("solid", start_color="7B3F00"),
        "h5"   : PatternFill("solid", start_color="005C5C"),
        "h6"   : PatternFill("solid", start_color="4A4A00"),
        "ok"   : PatternFill("solid", start_color="E8F5E9"),
        "warn" : PatternFill("solid", start_color="FFF3E0"),
        "miss" : PatternFill("solid", start_color="FCE4EC"),
        "no01" : PatternFill("solid", start_color="EDE7F6"),
        "zero" : PatternFill("solid", start_color="FFF9C4"),
        "excl" : PatternFill("solid", start_color="EEEEEE"),
        "total": PatternFill("solid", start_color="D6E4F0"),
    }


def _c(ws, row, col, val, s, align="c", fill=None, fkey="bf"):
    cell = ws.cell(row=row, column=col, value=val)
    na   = str(val) in ("-", "없음", "")
    cell.font      = s["gray"] if na else s[fkey]
    cell.border    = s["bd"]
    cell.alignment = s[align]
    if fill:
        cell.fill = fill
    return cell


def _hrow(ws, row, cols_fk, s):
    for col, (label, fk) in enumerate(cols_fk, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = s["hf"]; cell.fill = s[fk]
        cell.alignment = s["c"]; cell.border = s["bd"]


def _rfill(r, s):
    n = r["note"]
    if n.startswith("✅"):  return s["ok"]
    if n.startswith("⚠️"): return s["warn"]
    return s["miss"]


def _write_summary_sheet(ws, results, excluded, base_path: Path, prefix: str, s):
    """시트1: 종합 점검"""
    ws.merge_cells("A1:Z1")
    ws["A1"] = f"평가파일 종합 점검  |  prefix: {prefix}  |  {base_path}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12)
    ws["A1"].alignment = s["l"]
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:Z2")
    ws["A2"] = (
        f"생성일시: {datetime.now().strftime(FMT_DATETIME)}"
        f"  |  처리: {len(results)}건 / 제외: {len(excluded)}건"
    )
    ws["A2"].font = Font(name="Arial", size=9, color="777777")

    ws.append([])
    _hrow(ws, 4, [
        ("No", "h0"), ("폴더명", "h0"), ("파일 수", "h0"),
        ("변동 시트", "h1"), ("변동 B1", "h1"), ("Prefix 일치", "h1"),
        ("대장 시트", "h2"), ("대장 시트명", "h2"), ("미지급분배금", "h2"),
        ("F_1 위치", "h3"), ("펀드코드", "h3"), ("코드 유효", "h3"),
        ("_nav 위치", "h4"), ("_inv 위치", "h4"),
        ("장부기준일 셀", "h5"), ("장부기준일", "h5"), ("월일 일치", "h5"),
        ("미수이자A", "h5"), ("미수이자B", "h5"), ("미수이자", "h5"),
        ("기타자산A", "h5"), ("기타자산B", "h5"), ("기타자산 일치", "h5"),
        ("02.증빙 메일", "h6"), ("01.자료 수", "h6"),
        ("종합 상태", "h0"),
    ], s)

    for i, r in enumerate(results, 1):
        ri       = ws.max_row + 1
        fill     = _rfill(r, s)
        cnt01    = len(r["files_01"]) if isinstance(r["files_01"], list) else r["files_01"]
        f01_fill = s["no01"] if r["files_01"] == "해당없음" else fill

        _c(ws, ri,  1, i,                               s,        fill=fill)
        _c(ws, ri,  2, r["folder"],                     s, "l",   fill=fill)
        _c(ws, ri,  3, r["count"],                      s,        fill=fill)
        _c(ws, ri,  4, r["change_note"],                s, "l",   fill=fill)
        _c(ws, ri,  5, r["change_b1"]    or "-",        s, "l",   fill=fill)
        _c(ws, ri,  6, r["prefix_match"],               s,        fill=fill)
        _c(ws, ri,  7, r["daejang_exists"],              s,        fill=fill)
        _c(ws, ri,  8, r["daejang_name"] or "-",        s, "l",   fill=fill)
        _c(ws, ri,  9, r["daejang_kw"],                  s,        fill=fill)
        _c(ws, ri, 10, r["f1_addr"]      or "-",        s,        fill=fill)
        _c(ws, ri, 11, r["fund_code"]    or "-",        s,        fill=fill)
        _c(ws, ri, 12, "✅" if r["is_valid"] else "-",  s,        fill=fill)
        _c(ws, ri, 13, r["nav_ref"],                    s, "l",   fill=fill)
        _c(ws, ri, 14, r["inv_ref"],                    s, "l",   fill=fill)
        _c(ws, ri, 15, r["jangbu_cell"]  or "-",        s,        fill=fill)
        jb = r["jangbu_value"]
        _c(ws, ri, 16, jb if jb else "-",               s,        fill=fill)
        ws.cell(ri, 16).number_format = "@"
        _c(ws, ri, 17, r["date_match"],                 s,        fill=fill)
        _c(ws, ri, 18, r["misui_a"] if r["misui_a"] != "" else "-", s, fill=fill)
        _c(ws, ri, 19, r["misui_b"] if r["misui_b"] != "" else "-", s, fill=fill)
        _c(ws, ri, 20, r["misui_note"],                 s,        fill=fill)
        _c(ws, ri, 21, r["gita_a"]  if r["gita_a"]  != "" else "-", s, fill=fill)
        _c(ws, ri, 22, r["gita_b"]  if r["gita_b"]  != "" else "-", s, fill=fill)
        _c(ws, ri, 23, r["gita_note"],                  s,        fill=fill)
        _c(ws, ri, 24, r["mail_note"],                  s,        fill=fill)
        _c(ws, ri, 25, cnt01,                           s,        fill=f01_fill)
        _c(ws, ri, 26, r["note"],                       s, "l",   fill=fill)
        ws.row_dimensions[ri].height = 14

    tr   = ws.max_row + 1
    ok   = sum(1 for r in results if r["note"].startswith("✅"))
    warn = sum(1 for r in results if r["note"].startswith("⚠️"))
    ng   = len(results) - ok - warn

    _c(ws, tr,  1, "합계",                                        s, fkey="bold", fill=s["total"])
    _c(ws, tr,  2, f"처리 {len(results)} / 제외 {len(excluded)}", s, "l", fkey="bold", fill=s["total"])
    for col in range(3, 25):
        _c(ws, tr, col, "", s, fill=s["total"])
    _c(ws, tr, 24, f"메일없음: {sum(1 for r in results if '❌' in r['mail_note'])}건", s, fkey="bold", fill=s["total"])
    _c(ws, tr, 25, f"자료없음: {sum(1 for r in results if r['files_01'] == '해당없음')}건", s, fkey="bold", fill=s["total"])
    _c(ws, tr, 26, f"✅{ok}  ⚠️{warn}  ❌{ng}", s, fkey="bold", fill=s["total"])

    col_widths = [5, 44, 8, 20, 14, 12, 10, 12, 14, 18, 12, 10,
                  22, 22, 18, 14, 10, 14, 14, 14, 14, 14, 14, 14, 12, 14]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 16
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def _write_filelist_sheet(ws, results, s):
    """시트2: 01.자료 목록"""
    _hrow(ws, 1, [("폴더명", "h0"), ("파일 수", "h0"), ("파일명", "h0")], s)
    ri = 2
    for r in results:
        f01 = r["files_01"]
        if f01 == "해당없음":
            _c(ws, ri, 1, r["folder"], s, "l", fill=s["no01"])
            _c(ws, ri, 2, "해당없음",  s,       fill=s["no01"])
            _c(ws, ri, 3, "",          s,       fill=s["no01"])
            ri += 1
        elif len(f01) == 0:
            _c(ws, ri, 1, r["folder"], s, "l", fill=s["zero"])
            _c(ws, ri, 2, 0,           s,       fill=s["zero"])
            _c(ws, ri, 3, "(빈 폴더)", s,       fill=s["zero"])
            ri += 1
        else:
            for j, fname in enumerate(f01):
                _c(ws, ri, 1, r["folder"] if j == 0 else "", s, "l")
                _c(ws, ri, 2, len(f01)   if j == 0 else "", s)
                _c(ws, ri, 3, fname,                          s, "l")
                ri += 1
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60
    ws.sheet_view.showGridLines = False


def _write_excluded_sheet(ws, excluded, s):
    """시트3: 제외 목록"""
    _hrow(ws, 1, [("폴더명", "h0"), ("제외 사유", "h0")], s)
    for i, ex in enumerate(excluded, 2):
        _c(ws, i, 1, ex["folder"], s, "l", fill=s["excl"])
        _c(ws, i, 2, ex["reason"], s, "l", fill=s["excl"])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40
    ws.sheet_view.showGridLines = False


def _export(results, excluded, output_path: Path, base_path: Path, prefix: str):
    s  = _build_styles()
    wb = Workbook()

    ws = wb.active
    ws.title = "종합 점검"
    _write_summary_sheet(ws, results, excluded, base_path, prefix, s)

    _write_filelist_sheet(wb.create_sheet("01.자료"), results, s)
    _write_excluded_sheet(wb.create_sheet("제외 목록"), excluded, s)

    wb.save(output_path)


# ── Worker ────────────────────────────────────────────────────────────

class EvalValidatorWorker(AbstractWorker):
    """평가파일 종합 점검 — 폴더별 엑셀 파일 7종 검사 후 결과 저장"""

    def __init__(self, base_path: Path, prefix: str, include_ga: bool = False):
        super().__init__()
        self.base_path  = base_path
        self.prefix     = prefix.strip()
        self.include_ga = include_ga
        self._folders: List[Path] = []
        self._results  = []
        self._excluded = []

    @staticmethod
    def detect_prefixes(base_path: Path) -> List[str]:
        """하위 폴더 엑셀 파일명(1단계)에서 prefix 자동 감지 — 외부 공개 API"""
        return _detect_prefixes(base_path)

    def validate_inputs(self) -> bool:
        if not self.base_path or not self.base_path.is_dir():
            self.finished_error.emit(Msg.fmt(Msg.INVALID_PATH, path=self.base_path))
            return False
        if not self.prefix:
            self.finished_error.emit(Msg.PREFIX_EMPTY)
            return False
        return True

    def prepare(self) -> bool:
        self._folders = sorted(d for d in self.base_path.iterdir() if d.is_dir())
        if not self._folders:
            self.finished_error.emit(Msg.NO_SUBFOLDERS)
            return False
        self.stats["total"] = len(self._folders)
        self.emit_message(f"📁 {len(self._folders)}개 폴더 스캔 시작 (prefix: {self.prefix})")
        return True

    def execute(self):
        for i, folder in enumerate(self._folders, 1):
            if not self.check_running():
                return

            # 제외 폴더 ([로 시작)
            if folder.name.startswith("["):
                reason = _extract_bracket_reason(folder.name)
                self._excluded.append({"folder": folder.name, "reason": reason})
                self.emit_message(f"  ⊘ 제외: {folder.name} ({reason})", "warning")
                self.emit_progress(i, self.stats["total"])
                continue

            files     = _find_target_files(folder, self.prefix, self.include_ga)
            files_01  = _get_01_filenames(folder)
            mail_note = _check_02_mails(folder, self.prefix)

            if not files:
                self._results.append({
                    "folder": folder.name, "count": 0, "files": [],
                    "change_note": "-", "change_b1": "", "prefix_match": "-",
                    "daejang_exists": "-", "daejang_name": "", "daejang_kw": "-",
                    "f1_addr": "", "target_cell": "", "fund_code": "", "is_valid": False,
                    "nav_ref": "-", "inv_ref": "-",
                    "jangbu_cell": "", "jangbu_value": "", "date_match": "-",
                    "misui_a": "", "misui_b": "", "misui_note": "-",
                    "gita_a": "", "gita_b": "", "gita_note": "-",
                    "mail_note": mail_note, "files_01": files_01,
                    "note": "❌ 파일 없음", "error": "",
                })
                self.emit_message(f"  · {folder.name} → 파일 없음 | 메일: {mail_note}", "warning")
                self.emit_progress(i, self.stats["total"])
                continue

            for fp in files:
                if not self.check_running():
                    return

                info = _analyze_file(fp, self.prefix)

                checks = [
                    info["change_note"].startswith("✅"),
                    info["prefix_match"] == "✅",
                    info["daejang_exists"] == "✅",
                    info["is_valid"],
                    info["nav_ref"] != "없음",
                    info["inv_ref"] != "없음",
                    info["date_match"] == "✅",
                    info["gita_note"] == "✅",
                    mail_note.startswith("✅"),
                ]
                warn_flags = [
                    info["change_note"].startswith("⚠️"),
                    info["prefix_match"] == "❌",
                    info["date_match"] == "❌",
                    info["misui_note"].startswith("⚠️"),
                ]

                if all(checks):       note = "✅ 정상"
                elif any(warn_flags): note = "⚠️ 확인 필요"
                else:                 note = "❌ 오류"

                self._results.append({
                    "folder": folder.name, "count": len(files),
                    "files": [f.name for f in files],
                    **{k: info[k] for k in info},
                    "mail_note": mail_note, "files_01": files_01, "note": note,
                })

                icon = "✔" if note.startswith("✅") else ("!" if note.startswith("⚠️") else "✗")
                self.emit_message(
                    f"  {icon} {folder.name} / {fp.name} → {note}",
                    "success" if note.startswith("✅") else
                    "warning" if note.startswith("⚠️") else "error"
                )

            self.emit_progress(i, self.stats["total"])

    def finalize(self):
        if not self._results and not self._excluded:
            return

        output_path = self.base_path / f"종합점검_{self.prefix}_{fmt_file_stamp()}.xlsx"
        try:
            _export(self._results, self._excluded, output_path, self.base_path, self.prefix)
        except Exception as e:
            self.finished_error.emit(Msg.fmt(Msg.EXCEL_SAVE_FAIL, error=e))
            return

        ok   = sum(1 for r in self._results if r["note"].startswith("✅"))
        warn = sum(1 for r in self._results if r["note"].startswith("⚠️"))
        ng   = len(self._results) - ok - warn

        self.emit_message("─" * 50)
        self.emit_message(f"정상: {ok} / 확인필요: {warn} / 오류: {ng} / 제외: {len(self._excluded)}")
        self.emit_message(f"📄 저장 → {output_path.name}", "success")
        self.finished_success.emit(
            f"점검 완료 — 정상: {ok} / 확인필요: {warn} / 오류: {ng}\n결과 파일: {output_path.name}"
        )
